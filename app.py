import csv
import os
import re
import time
from datetime import datetime, timedelta
from collections import defaultdict
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    jsonify,
    send_from_directory,
    abort,
)
from werkzeug.utils import secure_filename
import json
import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1 import field_path

app = Flask(__name__)


def _read_secret(name: str) -> str | None:
    path = os.path.join('/etc/secrets', name)
    if os.path.exists(path):
        with open(path) as f:
            return f.read().strip()
    return None


app.secret_key = _read_secret('flask_secret_key') or os.environ.get(
    'FLASK_SECRET_KEY', 'replace-me'
)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
FIELDNAMES = [
    'Name',
    'Date',
    'Duration',
    'Task',
    'Description',
    'File',
    'Created At',
]

# Firebase initialization
cert_path = os.path.join(
    '/etc/secrets', 'sharptimetracker-firebase-adminsdk-fbsvc-973348138e.json'
)
if os.path.exists(cert_path):
    cred = credentials.Certificate(cert_path)
    firebase_admin.initialize_app(cred)
else:
    cred_json = os.environ.get('FIREBASE_CERT')
    if cred_json:
        cred = credentials.Certificate(json.loads(cred_json))
        firebase_admin.initialize_app(cred)
    else:
        raise RuntimeError('Firebase credentials not found')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def _map_row(row: dict) -> dict:
    key_map = {
        'name': 'Name',
        'date': 'Date',
        'task': 'Task',
        'description': 'Description',
        'file': 'File',
        'duration': 'Duration',
        'created_at': 'Created At',
        'created at': 'Created At',
    }
    out = {}
    for k, v in row.items():
        kk = k.lower().replace(' ', '_')
        if kk in key_map:
            out[key_map[kk]] = v
        else:
            out[k] = v
    for f in FIELDNAMES:
        out.setdefault(f, '')
    return out

def _read_entries():
    docs = (
        firestore.client()
        .collection('time_entries')
        .order_by(field_path.FieldPath('Created At').to_api_repr(), direction=firestore.Query.DESCENDING)
        .stream()
    )
    entries = []
    for doc in docs:
        row = doc.to_dict()
        mapped = _map_row(row)
        mapped['id'] = doc.id
        mapped['Date'] = _canonical_date(mapped['Date'])
        entries.append(mapped)
    return entries

def _hours(duration: str) -> float:
    """Return hours as float based on a duration string."""
    return _parse_duration(duration)

def _hm(hours: float) -> str:
    hrs = int(hours)
    mins = int(round((hours - hrs) * 60))
    return f"{hrs}:{mins:02d}"


def _parse_duration(dur: str) -> float:
    """Parse a duration like '1:30' or '1.5' into hours."""
    if not dur:
        return 0.0
    dur = dur.strip().lower().replace('h', '').replace('hrs', '').replace('hours', '')
    if ':' in dur:
        h, m = dur.split(':', 1)
        try:
            return int(h) + int(m) / 60
        except ValueError:
            return 0.0
    try:
        return float(dur)
    except ValueError:
        return 0.0

def _parse_date(date_str):
    formats = ['%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(date_str)
    except ValueError:
        return datetime.now()

def _canonical_date(date_str):
    return _parse_date(date_str).strftime('%Y-%m-%d')

def _format_date(date_str, show_year=False):
    dt = _parse_date(date_str)
    return dt.strftime('%m/%d/%Y' if show_year else '%m/%d')

def _format_date_long(date_str):
    dt = _parse_date(date_str)
    return dt.strftime('%-d %B')

@app.template_filter('linkify')
def _linkify(text):
    url_pattern = r'(https?://[\w\.-/]+)'
    return re.sub(url_pattern, r'<a href="\1" target="_blank">\1</a>', text)

def _weekly_summary(entries):
    weekly = defaultdict(lambda: defaultdict(float))
    for row in entries:
        dt = _parse_date(row['Date'])
        week_start = dt - timedelta(days=dt.weekday())
        key = week_start.strftime('%Y-%m-%d')
        weekly[key][row['Name']] += _hours(row.get('Duration', ''))
    return weekly

def _daily_summary(entries):
    daily = defaultdict(lambda: defaultdict(float))
    for row in entries:
        key = _canonical_date(row['Date'])
        daily[key][row['Name']] += _hours(row.get('Duration', ''))
    return daily

def _weekday_summary(entries):
    days = defaultdict(lambda: defaultdict(float))
    labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for row in entries:
        dt = _parse_date(row['Date'])
        key = labels[dt.weekday()]
        days[key][row['Name']] += _hours(row.get('Duration', ''))
    return days

def _week_list(entries):
    weeks = sorted(_weekly_summary(entries).keys())
    return weeks


@app.route('/')
def index():
    entries = _read_entries()
    years = {_parse_date(e['Date']).year for e in entries}
    show_year = any(year != 2025 for year in years)
    for e in entries:
        e['hours'] = _hours(e.get('Duration', ''))
        hrs = int(e['hours'])
        mins = int(round((e['hours'] - hrs) * 60))
        e['duration_str'] = f"{hrs}h {mins}m"
        e['duration_hm'] = f"{hrs}:{mins:02d}"
        e['date_display'] = _format_date(e['Date'], show_year)

    grouped = defaultdict(list)
    totals = {}
    for e in entries:
        key = _canonical_date(e['Date'])
        grouped[key].append(e)
    for date, rows in grouped.items():
        per = defaultdict(float)
        for r in rows:
            per[r['Name']] += r['hours']
        totals[date] = per

    ordered = dict(sorted(grouped.items()))
    return render_template(
        'index.html',
        grouped=ordered,
        totals=totals,
        show_year=show_year,
        format_date=_format_date,
    )

@app.route('/add', methods=['POST'])
def add():
    today = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))
    today = _canonical_date(today)
    uploaded = request.files.get('file')
    filename = ''
    if uploaded and uploaded.filename:
        safe = secure_filename(uploaded.filename)
        fname = f"{int(time.time())}_{safe}"
        path = os.path.join(UPLOAD_FOLDER, fname)
        uploaded.save(path)
        filename = fname
    row = {
        'Name': request.form['name'],
        'Date': today,
        'Duration': request.form.get('duration', ''),
        'Task': request.form['task'],
        'Description': request.form.get('description', ''),
        'File': filename,
        'Created At': datetime.now().isoformat(),
    }
    doc_ref = firestore.client().collection('time_entries').document()
    doc_ref.set(row)
    doc_id = doc_ref.id

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        hours = _hours(row.get('Duration', ''))
        years = {_parse_date(r['Date']).year for r in _read_entries()}
        show_year = any(y != 2025 for y in years)
        return jsonify(
            {
                'name': row['Name'],
                'date_display': _format_date(row['Date'], show_year),
                'hours': hours,
                'duration_str': f"{int(hours)}h {int(round((hours-int(hours))*60))}m",
                'duration_hm': f"{int(hours)}:{int(round((hours-int(hours))*60)):02d}",
                'task': row['Task'],
                'description_html': _linkify(row['Description']),
                'file_link': f'<a href="/uploads/{filename}" download target="_blank">{filename}</a>' if filename else '',
                'id': doc_id,
            }
        )

    flash('Time entry added successfully')
    return redirect(url_for('index'))

@app.route('/import', methods=['POST'])
def import_csv():
    uploaded = request.files.get('csv')
    if not uploaded or not uploaded.filename:
        return jsonify({'error': 'no file'}), 400
    col = firestore.client().collection('time_entries')

    # build set of existing entries to avoid duplicates (date+duration+task)
    existing = set()
    for doc in col.stream():
        d = doc.to_dict()
        key = (
            _canonical_date(d.get('Date', '')),
            d.get('Duration', ''),
            d.get('Task', ''),
        )
        existing.add(key)

    seen = set()

    ext = os.path.splitext(uploaded.filename)[1].lower()
    rows = []
    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(uploaded.read()), data_only=True)
        ws = wb.active
        iter_rows = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(iter_rows, [])]
        for values in iter_rows:
            rows.append({headers[i]: (values[i] if values[i] is not None else '') for i in range(len(headers))})
    else:
        reader = csv.DictReader(uploaded.stream.read().decode('utf-8').splitlines())
        rows = list(reader)

    today = datetime.now().strftime('%Y-%m-%d')
    for row in rows:
        date = _canonical_date(row.get('Date', datetime.now().strftime('%Y-%m-%d')))
        duration = row.get('Duration', '')
        task = row.get('Task', '')
        key = (date, duration, task)
        if key in existing or key in seen:
            continue
        seen.add(key)
        data = {
            'Name': row.get('Name', ''),
            'Date': date,
            'Duration': duration,
            'Task': task,
            'Description': row.get('Description', ''),
            'File': row.get('File', ''),
            'Created At': datetime.now().isoformat(),
        }
        col.add(data)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True})
    flash('Data imported')
    return redirect(url_for('index'))

@app.route('/report')
def report():
    return render_template('report.html')

@app.route('/weekly-data')
def weekly_data():
    weekly = _weekly_summary(_read_entries())
    weeks = sorted(weekly.keys())
    years = {_parse_date(w).year for w in weeks}
    show_year = any(y != 2025 for y in years)
    labels = [_format_date(w, show_year) for w in weeks]
    names = sorted({name for w in weekly.values() for name in w.keys()})
    hours = {name: [weekly[w].get(name, 0) for w in weeks] for name in names}
    return jsonify({'weeks': labels, 'names': names, 'hours': hours})

@app.route('/daily-data')
def daily_data():
    daily = _daily_summary(_read_entries())
    dates = sorted(daily.keys())
    names = sorted({n for d in daily.values() for n in d.keys()})
    hours = {name: [daily[date].get(name, 0) for date in dates] for name in names}
    return jsonify({'dates': dates, 'names': names, 'hours': hours})

@app.route('/weekdays-data')
def weekdays_data():
    days = _weekday_summary(_read_entries())
    labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    names = sorted({n for d in days.values() for n in d.keys()})
    hours = {name: [days[lab].get(name, 0) for lab in labels] for name in names}
    return jsonify({'days': labels, 'names': names, 'hours': hours})

@app.route('/week-list')
def week_list():
    weeks = _week_list(_read_entries())
    return jsonify({'weeks': weeks})

@app.route('/week-data')
def week_data():
    start = request.args.get('start')
    if not start:
        return jsonify({'error': 'missing start'}), 400
    dt = _parse_date(start)
    days = [(dt + timedelta(days=i)) for i in range(7)]
    day_keys = [d.strftime('%Y-%m-%d') for d in days]
    labels = [_format_date_long(d.strftime('%Y-%m-%d')) for d in days]
    entries = [e for e in _read_entries() if _canonical_date(e['Date']) in day_keys]
    data = defaultdict(lambda: defaultdict(float))
    weekday_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for e in entries:
        d = _parse_date(e['Date']).weekday()
        key = weekday_labels[d]
        data[key][e['Name']] += _hours(e.get('Duration', ''))
    names = sorted({n for d in data.values() for n in d.keys()})
    hours = {name: [data[weekday_labels[i]].get(name, 0) for i in range(7)] for name in names}
    return jsonify({'days': labels, 'names': names, 'hours': hours})

@app.route('/download')
def download():
    entries = _read_entries()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Name',
        'Date',
        'Duration',
        'Task',
        'Description',
        'File',
    ])
    for e in entries:
        ws.append([
            e.get('Name', ''),
            e['Date'],
            _hm(_hours(e.get('Duration', ''))),
            e['Task'],
            e['Description'],
            e['File'],
        ])
    file = 'entries.xlsx'
    wb.save(file)
    return send_file(
        file,
        as_attachment=True,
        download_name='Sharp Time Tracker.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@app.route('/weekly-download')
def weekly_download():
    weekly = _weekly_summary(_read_entries())
    weeks = sorted(weekly.keys())
    years = {_parse_date(w).year for w in weeks}
    show_year = any(y != 2025 for y in years)
    labels = [_format_date(w, show_year) for w in weeks]
    names = sorted({name for w in weekly.values() for name in w.keys()})
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Week'] + names)
    for week, label in zip(weeks, labels):
        ws.append([label] + [weekly[week].get(name, 0) for name in names])
    file = 'weekly_report.xlsx'
    wb.save(file)
    return send_file(
        file,
        as_attachment=True,
        download_name='weekly_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@app.route('/uploads/<path:filename>')
def uploaded(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@app.route('/edit/<entry_id>', methods=['GET', 'POST'])
def edit(entry_id):
    doc_ref = firestore.client().collection('time_entries').document(entry_id)
    snap = doc_ref.get()
    if not snap.exists:
        abort(404)
    row = _map_row(snap.to_dict())
    created_str = row.get('Created At', datetime.now().isoformat())
    if not isinstance(created_str, str):
        created_str = str(created_str)
    try:
        created = datetime.fromisoformat(created_str)
    except ValueError:
        created = datetime.now()

    if datetime.now() - created > timedelta(hours=24):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'expired'}), 400
        flash('Editing period expired')
        return redirect(url_for('index'))
    if request.method == 'POST':
        data = {
            'Date': _canonical_date(request.form['date']),
            'Duration': request.form.get('duration', ''),
            'Task': request.form['task'],
            'Description': request.form.get('description', ''),
        }
        doc_ref.update(data)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True})
        flash('Entry updated')
        return redirect(url_for('index'))
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = dict(row)
        data['id'] = entry_id
        return jsonify(data)
    return render_template('edit.html', row=row, index=entry_id)

@app.route('/delete/<entry_id>', methods=['POST'])
def delete(entry_id):
    firestore.client().collection('time_entries').document(entry_id).delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True})
    flash('Entry deleted')
    return redirect(url_for('index'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
